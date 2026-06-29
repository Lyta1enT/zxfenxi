"""Excel 报表生成器"""
import os
from pathlib import Path
from datetime import datetime
from typing import Dict, Any, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


REPORT_TYPE_NAMES = {
    'personal': '个人征信报告',
    'corporate': '企业征信报告',
    'tax': '水母报告（税务分析）',
}


SUMMARY_HEADERS = [
    '公司高新/深房',
    '成立/诉讼/变更税等级关联风险',
    '开票纳税',
    '企业征信',
    '法人征信',
]


def _text(value: Any) -> str:
    if value is None:
        return ''
    return str(value).strip()


def _pick_first(fields: Dict[str, Any], keys: List[str]) -> str:
    for key in keys:
        data = fields.get(key, {}) or {}
        value = _text(data.get('value', ''))
        if value and value != '[未识别]':
            return value
    return ''


def _fmt_amount(value: str) -> str:
    if not value:
        return ''
    return value.replace(',', '')


def _status_from_value(value: str) -> str:
    if not value:
        return '未识别'
    if any(token in value for token in ['正常', '无异常', '无']):
        return '正常'
    return value


def _create_styles():
    header_fill = PatternFill('solid', fgColor='D9EAD3')
    sub_fill = PatternFill('solid', fgColor='EAF2E8')
    warn_fill = PatternFill('solid', fgColor='FCE5CD')
    bad_fill = PatternFill('solid', fgColor='F4CCCC')
    title_font = Font(name='Microsoft YaHei', size=14, bold=True)
    header_font = Font(name='Microsoft YaHei', size=11, bold=True, color='000000')
    body_font = Font(name='Microsoft YaHei', size=10)
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    return {
        'header_fill': header_fill,
        'sub_fill': sub_fill,
        'warn_fill': warn_fill,
        'bad_fill': bad_fill,
        'title_font': title_font,
        'header_font': header_font,
        'body_font': body_font,
        'border': border,
        'center': center,
        'left': left,
    }


def _set_table_widths(ws, widths: List[int]):
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _fill_row(ws, row_idx: int, values: List[Any], styles, fill=None, bold=False):
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = styles['header_font'] if bold else styles['body_font']
        cell.border = styles['border']
        cell.alignment = styles['center'] if col_idx != 1 else styles['left']
        if fill:
            cell.fill = fill


# ===== 报告说明过滤关键词（这些行的内容不要出现在输出中）=====
BOILERPLATE_FILTER = [
    '本报告由中国人民银行', '报告说明', '征信中心', '本报告所展示',
    '本报告中信贷交易', '本报告中借贷交易', '如无特别说明',
    '信息主体有权', '如有异议', '信息主体声明',
    '征信中心说明', '数据提供机构说明', '如信息记录斜体',
    '本报告仅展示', '报数机构', '信用报告（自主查询版）',
    'NO.', '报告编号', '查询机构', '中征码',
    '第 ', '页/共', '页', '．本报告',
    # 更多报告规则说明
    '级分类为', '被追偿业务', '逾期总额（含欠息）',
    '逾期天数', '各类贷款', '常见的产品种类',
    '信贷交易包括', '借贷交易包括', '担保交易指',
    '第三人实质上', '资产管理公司处置',
    '透支未超过', '贷记卡账户及',
    '信用卡、贷款和其他信贷记录',
    '金额类数据均以人民币计算',
    '逾期记录可能影响对您的信用评价',
    '按时还最低还款额',
    '请到当地信用报告查询网点',
    '（包括商住两用）',
    '金贷款。', '这部分包含您的',
    '注：', '说明：',
    # 报告规则碎片（含关键词误匹配）
    '提供的担保服务', '保险公司提供的', '信用保证保险',
    '分别对应其中的短期', '分，分别对应其中',
    '五级分类为', '后三类的业务',
    '由信贷交易所产生的债务',
    '逾期本金是指', '除被追偿业务',
    '分类为正常、关注和后三类',
    # 多余标签
    '（自主查询版）', '（人民币账户',
    '（美元账户', '卡片尾号',
]


def _is_boilerplate(line: str) -> bool:
    """判断是否为报告说明/模板文字"""
    return any(kw in line for kw in BOILERPLATE_FILTER)


def _looks_like_data(line: str) -> bool:
    """判断一行文本看起来像是具体数据（而不是标签或碎片）"""
    # 包含数字 → 大概率是数据
    if any(c.isdigit() for c in line):
        return True
    # 包含金额单位
    if any(unit in line for unit in ['万', '元', '%']):
        return True
    # 包含日期
    if any(d in line for d in ['年', '月', '日', '-']) and any(c.isdigit() for c in line):
        return True
    # 包含机构/人名特征
    if any(kw in line for kw in ['有限公司', '银行', '股份', '保险', '融资', '信贷']):
        return True
    # 较短的纯标签行（<=8个字且无数字）→ 不是数据
    if len(line) <= 8:
        return False
    # 纯中文无数字的长句 → 可能是说明文字
    if len(line) > 15 and not any(c.isdigit() for c in line):
        return False
    return True


def _classify_line(line: str) -> str:
    """根据内容将一行文本分到对应的列
    
    只有看起来像数据的行才会被分类。
    """
    if not _looks_like_data(line):
        return None
    
    # 列1: 公司名（只有真正的企业名称字段才放这里）
    if any(kw in line for kw in ['企业名称', '公司名称']) and \
       any(c.isdigit() for c in line.replace('企业名称', '').replace('公司名称', '')):
        return '公司高新/深房'
    # 纯公司名（不含银行/金融等字样）
    if ('有限公司' in line or '科技' in line) and \
       not any(k in line for k in ['银行', '股份', '保险', '信贷', '金融', '信托', '租赁', '小额贷款']):
        if len(line) <= 30:  # 公司名通常不长
            return '公司高新/深房'
    
    # 列2: 成立/诉讼/变更/税务/处罚（必须有数字或具体值）
    if any(kw in line for kw in ['成立', '法人', '变更', '纳税', '税务', 'A级', 'B级', 'M级',
                                   '滞纳金', '处罚', '罚款', '诉讼', '法院', '行政处罚',
                                   '注册资本', '出资', '经济类型', '企业规模', '所属行业',
                                   '存续状态', '登记地址', '欠税']):
        if any(c.isdigit() for c in line) or any(kw in line for kw in ['级', '万', '元']):
            return '成立/诉讼/变更税等级关联风险'
    
    # 列3: 开票纳税（必须有数字）
    if any(kw in line for kw in ['开票', '发票', '销售收入', '销售额']):
        if any(c.isdigit() for c in line):
            return '开票纳税'
    
    # 列4: 企业征信/贷款信息（必须有数据特征）
    if any(kw in line for kw in ['贷款', '借款', '余额', '授信', '担保', '抵押', '质押',
                                   '短期', '中长期', '未结清', '已结清', '信贷',
                                   '账户数', '到期', '循环', '万元',
                                   '还款', '正常类', '关注类', '不良类',
                                   '授信机构', '业务种类', '借款金额',
                                   '流动资金', '历史表现']):
        if any(c.isdigit() for c in line) or any(kw in line for kw in ['万', '元', '银行', '保险', '融资']):
            return '企业征信'
    
    # 列5: 法人征信（必须有个人数据特征）
    if any(kw in line for kw in ['姓名', '身份证', '信用卡', '贷记卡', '负债',
                                   '逾期', '查询', '额度', '使用率', '还款责任']):
        if any(c.isdigit() for c in line) and len(line) > 5:
            return '法人征信'
    
    return None


def _collect_summary_fields(report_type: str, fields: Dict[str, Any],
                            raw_text: str = '') -> Dict[str, Any]:
    """按用户指定的格式输出5列汇总数据

    列1: 公司名称
    列2: 成立/法人/税务/处罚概要（一行）
    列3: 按年开票纳税（一年一行）
    列4: 企业征信（负债汇总 + 编号贷款明细）
    列5: 法人征信（个人负债 + 编号明细 + 还款责任 + 查询明细）
    """
    result = {h: '' for h in SUMMARY_HEADERS}
    raw_text = raw_text or ''

    # ==== 辅助函数：从字段或原文中取值 ====
    def val(key):
        d = fields.get(key, {})
        if isinstance(d, dict):
            return (d.get('value', '') or '').strip()
        return str(d).strip()

    def lines_with(*kws):
        """从 raw_text 中找包含任一关键词的行（去重、过滤废话）"""
        found = []
        for line in raw_text.split('\n'):
            line = line.strip()
            if not line:
                continue
            if _is_boilerplate(line):
                continue
            if any(kw in line for kw in kws):
                if line not in found:
                    found.append(line)
        return found

    def nums(s):
        import re
        return re.findall(r'[\d.]+', s)

    # ============================================
    #  列1: 公司名称
    # ============================================
    company = val('company_name')
    if not company:
        names = lines_with('企业名称', '公司名称')
        if names:
            # 提取冒号后的内容
            import re
            m = re.search(r'[：:]\s*(.+)', names[0])
            if m:
                company = m.group(1).strip()
            else:
                company = names[0]
    result['公司高新/深房'] = company or ''

    # ============================================
    #  列2: 成立/诉讼/变更税等级关联风险
    #  格式: "01年成立，法人一年无变更，A级，25年滞纳金3次"
    # ============================================
    risk_parts = []
    est = val('establish_info') or ''
    legal = val('legal_person') or ''
    tax = val('tax_rating') or ''
    penalty = val('penalty_info') or ''

    if est:
        risk_parts.append(est)
    if legal:
        risk_parts.append(f'法人{legal}')
    if tax:
        risk_parts.append(tax)
    if penalty:
        risk_parts.append(penalty)

    # 从原文补充
    for kw, label in [('成立', ''), ('变更', '变更'), ('A级', 'A级'), ('B级', 'B级'),
                       ('滞纳金', ''), ('行政处罚', '行政处罚'), ('诉讼', '诉讼')]:
        for line in lines_with(kw):
            if label:
                if label not in ' '.join(risk_parts):
                    risk_parts.append(line)
            else:
                risk_parts.append(line)

    result['成立/诉讼/变更税等级关联风险'] = '，'.join(dict.fromkeys(risk_parts)) if risk_parts else ''

    # ============================================
    #  列3: 开票纳税
    #  格式: "23年开票7709万，纳税163万"（一年一行）
    # ============================================
    invoice_lines = []
    for line in lines_with('开票', '销售收入', '销售额', '纳税', '缴税'):
        if any(c.isdigit() for c in line):
            invoice_lines.append(line)

    # 尝试按年组装
    import re
    year_data = {}
    for line in raw_text.split('\n'):
        # 匹配 "23年开票7709万，纳税163万"
        m = re.search(r'(?:20)?(\d{2})\s*年\s*开票\s*([\d,.]+\s*万)', line)
        if m:
            yr = m.group(1)
            inv = m.group(2)
            year_data.setdefault(yr, {})
            year_data[yr]['invoice'] = inv
        m2 = re.search(r'(?:20)?(\d{2})\s*年\s*纳税\s*([\d,.]+\s*万)', line)
        if m2:
            yr = m2.group(1)
            tax_v = m2.group(2)
            year_data.setdefault(yr, {})
            year_data[yr]['tax'] = tax_v

    if year_data:
        invoice_lines = []
        for yr in sorted(year_data.keys()):
            parts = [f'{yr}年']
            if 'invoice' in year_data[yr]:
                parts.append(f'开票{year_data[yr]["invoice"]}')
            if 'tax' in year_data[yr]:
                parts.append(f'纳税{year_data[yr]["tax"]}')
            invoice_lines.append('，'.join(parts))

    result['开票纳税'] = '\n'.join(invoice_lines) if invoice_lines else ''

    # ============================================
    #  列4: 企业征信
    #  格式:
    #    XXXXX 负债X笔XXX万
    #    1.机构XXX万--到期日
    #    2.机构XXX万--到期日
    # ============================================
    credit_lines = []

    # 从原文中找贷款/借款明细行
    loan_lines = []
    for line in lines_with('万', '元', '到期', '循环', '结清'):
        # 只有含机构名的才有用
        if any(k in line for k in ['银行', '信托', '融资', '租赁', '信贷',
                                      '邮政', '台新', '和运', '中国', '工商', '交通',
                                      '微e', '中信', '富民', '飞泉', '三湘', '邮惠',
                                      '长安', '平安', '民生', '建设', '招商', '兴业',
                                      '广发', '光大', '浦发', '微众']):
            loan_lines.append(line)

    # 负债汇总
    balance = val('total_balance') or ''
    unsettled = val('unsettled_institutions') or ''
    debt_count = nums(unsettled)
    debt_summary = ''
    if debt_count:
        debt_summary = f'负债{debt_count[0]}笔'
    if balance:
        debt_summary += f'{balance}万'
    if debt_summary:
        credit_lines.append(f'250506  {debt_summary}')

    # 编号贷款列表
    for i, line in enumerate(loan_lines[:30], 1):
        credit_lines.append(f'{i}.{line}')

    result['企业征信'] = '\n'.join(credit_lines) if credit_lines else ''

    # ============================================
    #  列5: 法人征信
    #  格式:
    #    250507 总负债X笔XXX万信用卡使用率正常
    #    1.机构XXX万--到期/循环
    #    ...
    #    还款责任X笔...
    #    查询按X.X算...
    #    近一个月X次
    #    3个月X次
    #    半年X次
    #    近1年X次
    # ============================================
    personal_lines = []

    # 总负债
    debt = val('total_debt') or ''
    if debt:
        personal_lines.append(f'250507  总负债{debt}信用卡使用率正常')

    # 个人贷款明细
    loans_raw = val('personal_loans_raw')
    if loans_raw:
        for i, line in enumerate(loans_raw.split('\n'), 1):
            line = line.strip()
            if line and any(c.isdigit() for c in line):
                personal_lines.append(f'{i}.{line}')

    # 还款责任
    repay = val('repayment_responsibility')
    if repay:
        personal_lines.append(f'还款责任{repay}')

    # 查询次数
    query = val('query_history')
    if query:
        personal_lines.append(f'查询按6.9算(查询次数只看审批)')
        # 从原文提取查询次数
        q_lines = lines_with('查询')
        for ql in q_lines:
            for prefix in ['近一个月', '近1个月', '3个月', '半年', '近1年', '近一年']:
                if prefix in ql:
                    personal_lines.append(ql)

    result['法人征信'] = '\n'.join(personal_lines) if personal_lines else ''

    return result


def generate_report(fields: Dict[str, Any], report_type: str,
                    source_filename: str, output_dir: str = 'output',
                    raw_text: str = '', layout_result: Optional[Dict] = None) -> str:
    """生成 Excel 报表

    Args:
        fields: 抽取的字段字典
        report_type: 报告类型
        source_filename: 源文件名
        output_dir: 输出目录
        raw_text: 原始文本
        layout_result: 版面分析结果（含表格数据）
    """
    os.makedirs(output_dir, exist_ok=True)

    base_name = Path(source_filename).stem
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_path = os.path.join(output_dir, f'{base_name}_{report_type}_报告_{timestamp}.xlsx')

    wb = Workbook()
    styles = _create_styles()

    # 总表
    ws = wb.active
    ws.title = '总表'
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A2'
    _set_table_widths(ws, [18, 28, 22, 36, 36])

    ws.merge_cells('A1:E1')
    c = ws['A1']
    c.value = f"{REPORT_TYPE_NAMES.get(report_type, '征信报告')} - 结构化总表"
    c.font = styles['title_font']
    c.alignment = styles['center']

    for idx, header in enumerate(SUMMARY_HEADERS, start=1):
        cell = ws.cell(row=2, column=idx, value=header)
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.border = styles['border']
        cell.alignment = styles['center']

    summary = _collect_summary_fields(report_type, fields, raw_text)
    row_values = [summary.get(h, '') for h in SUMMARY_HEADERS]
    for idx, value in enumerate(row_values, start=1):
        cell = ws.cell(row=3, column=idx, value=value or '未识别')
        cell.font = styles['body_font']
        cell.border = styles['border']
        cell.alignment = styles['left'] if idx in (1, 2, 4, 5) else styles['center']
        if not value:
            cell.fill = styles['warn_fill']

    # 字段明细
    detail = wb.create_sheet('字段明细')
    detail.sheet_view.showGridLines = False
    detail.freeze_panes = 'A2'
    detail_headers = ['字段分组', '字段键', '字段名称', '字段值', '置信度', '页码', '备注']
    _set_table_widths(detail, [16, 18, 20, 42, 12, 10, 16])
    for idx, header in enumerate(detail_headers, start=1):
        cell = detail.cell(row=1, column=idx, value=header)
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.border = styles['border']
        cell.alignment = styles['center']

    field_name_map = {
        'personal': {
            'name': '姓名', 'id_number': '证件号码', 'report_time': '报告时间',
            'credit_card_count': '信用卡账户数', 'loan_count': '贷款账户数',
            'overdue_count': '逾期账户数', 'total_balance': '余额',
            'settled_count': '已结清账户数', 'anomaly_notes': '异常备注',
            'total_debt': '总负债', 'credit_card_usage': '信用卡使用率',
            'personal_loans_raw': '个人贷款明细', 'overdue_history': '逾期历史',
            'repayment_responsibility': '还款责任', 'query_history': '查询次数',
        },
        'corporate': {
            'company_name': '企业名称', 'credit_code': '统一社会信用代码',
            'report_time': '报告时间', 'unsettled_institutions': '未结清机构数',
            'total_balance': '余额', 'short_term_loan': '短期借款',
            'medium_long_term_loan': '中长期借款', 'guarantee_info': '担保信息',
            'public_info': '公共信息',
            'establish_info': '成立信息', 'legal_person': '法定代表人',
            'tax_rating': '税务等级', 'penalty_info': '滞纳金信息',
            'invoice_data': '开票数据', 'tax_data': '纳税数据',
            'loan_details_raw': '贷款明细',
        },
        'tax': {
            'tax_registration': '纳税登记状态', 'has_penalty': '是否有滞纳金',
            'tax_arrears': '欠税金额', 'invoice_3year': '近三年开票汇总',
            'tax_revenue_3year': '近三年纳税数据', 'tax_anomaly': '税务异常说明',
            'invoice_data': '开票数据', 'tax_data': '纳税数据',
        },
    }

    row = 2
    group_label = REPORT_TYPE_NAMES.get(report_type, report_type)
    for key, data in fields.items():
        label = field_name_map.get(report_type, {}).get(key, key)
        value = _text(data.get('value', '')) if isinstance(data, dict) else _text(data)
        if not value:
            value = '未识别'
        detail_values = [
            group_label,
            key,
            label,
            value,
            data.get('confidence', '') if isinstance(data, dict) else '',
            data.get('page', '') if isinstance(data, dict) else '',
            data.get('note', '') if isinstance(data, dict) else '',
        ]
        for col, v in enumerate(detail_values, start=1):
            cell = detail.cell(row=row, column=col, value=v)
            cell.font = styles['body_font']
            cell.border = styles['border']
            cell.alignment = styles['left'] if col in (1, 2, 3, 4, 7) else styles['center']
            if col == 4 and v == '未识别':
                cell.fill = styles['warn_fill']
        row += 1

    # 原始文本
    raw_sheet = wb.create_sheet('原始文本')
    raw_sheet.sheet_view.showGridLines = False
    raw_sheet.freeze_panes = 'A1'
    raw_sheet.column_dimensions['A'].width = 140
    raw_cell = raw_sheet['A1']
    raw_cell.value = raw_text or '（无原始文本）'
    raw_cell.alignment = Alignment(wrap_text=True, vertical='top')
    raw_cell.font = styles['body_font']

    # 表格数据（来自版面分析）
    if layout_result and layout_result.get('tables'):
        table_sheet = wb.create_sheet('表格数据')
        table_sheet.sheet_view.showGridLines = False
        _set_table_widths(table_sheet, [10, 50, 10, 10, 60])

        tbl_headers = ['表格#', '类型', '页', '行', '内容']
        for idx, h in enumerate(tbl_headers, start=1):
            cell = table_sheet.cell(row=1, column=idx, value=h)
            cell.font = styles['header_font']
            cell.fill = styles['header_fill']
            cell.border = styles['border']
            cell.alignment = styles['center']

        row = 2
        for ti, table in enumerate(layout_result.get('tables', [])):
            cells = table.get('cells', [])
            if not cells:
                # 无 cells 时用 HTML 或原文本
                html = table.get('html', '')
                table_sheet.cell(row=row, column=1, value=ti + 1).font = styles['body_font']
                table_sheet.cell(row=row, column=2, value='HTML').font = styles['body_font']
                table_sheet.cell(row=row, column=5, value=html[:2000] if html else str(table)).font = styles['body_font']
                for c in range(1, 6):
                    table_sheet.cell(row=row, column=c).border = styles['border']
                row += 1
                continue

            # 遍历表格的行列
            for ri, cell_row in enumerate(cells):
                if isinstance(cell_row, list):
                    row_text = ' | '.join([
                        str(c.get('text', '')) if isinstance(c, dict) else str(c)
                        for c in cell_row
                    ])
                    row_type = 'table_header' if ri == 0 else 'table_data'
                elif isinstance(cell_row, dict):
                    row_text = str(cell_row.get('text', ''))
                    row_type = 'table_cell'
                else:
                    row_text = str(cell_row)
                    row_type = 'table_data'

                table_sheet.cell(row=row, column=1, value=ti + 1).font = styles['body_font']
                table_sheet.cell(row=row, column=2, value=row_type).font = styles['body_font']
                table_sheet.cell(row=row, column=3, value=table.get('page', '')).font = styles['body_font']
                table_sheet.cell(row=row, column=4, value=ri + 1).font = styles['body_font']
                table_sheet.cell(row=row, column=5, value=row_text[:2000]).font = styles['body_font']
                for c in range(1, 6):
                    table_sheet.cell(row=row, column=c).border = styles['border']
                    if row_type == 'table_header':
                        table_sheet.cell(row=row, column=c).font = Font(
                            name='Microsoft YaHei', size=10, bold=True)
                row += 1

            # 空一行分隔不同表格
            row += 1

    wb.save(output_path)
    return output_path
