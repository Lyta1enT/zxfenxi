"""文件类型判断和处理"""
import os
from pathlib import Path
from typing import Dict, Any, Optional

from app.utils.pdf_utils import extract_text_from_pdf_with_fallback
from app.utils.image_utils import preprocess_image


SUPPORTED_EXTENSIONS = {
    '.pdf': 'PDF',
    '.png': 'IMAGE',
    '.jpg': 'IMAGE',
    '.jpeg': 'IMAGE',
    '.bmp': 'IMAGE',
    '.tiff': 'IMAGE',
    '.docx': 'WORD',
    '.xlsx': 'EXCEL',
}


def get_file_type(file_path: str) -> str:
    """判断文件类型，返回类型代码"""
    ext = Path(file_path).suffix.lower()
    return SUPPORTED_EXTENSIONS.get(ext, 'UNKNOWN')


def is_supported(file_path: str) -> bool:
    """检查文件是否受支持"""
    return get_file_type(file_path) != 'UNKNOWN'


def process_file(file_path: str) -> Dict[str, Any]:
    """处理上传的文件，返回统一格式的结果
    
    Returns:
        {
            'file_type': str,
            'file_path': str,
            'text': str,          # 直接提取的文本（如有）
            'image_paths': list,   # 图片路径列表（用于OCR）
            'metadata': dict       # 文件元信息
        }
    """
    file_type = get_file_type(file_path)
    
    result = {
        'file_type': file_type,
        'file_path': file_path,
        'text': '',
        'image_paths': [],
        'metadata': {'pages': 0},
    }
    
    if file_type == 'PDF':
        text, images = extract_text_from_pdf_with_fallback(file_path)
        result['text'] = text
        result['image_paths'] = images
        result['metadata']['pages'] = len(images) if images else 1
    
    elif file_type == 'IMAGE':
        result['image_paths'] = [file_path]
        result['metadata']['pages'] = 1
    
    elif file_type == 'WORD':
        try:
            from docx import Document
            doc = Document(file_path)
            text = '\n'.join([p.text for p in doc.paragraphs])
            result['text'] = text
        except Exception:
            pass
    
    elif file_type == 'EXCEL':
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True)
            texts = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row in ws.iter_rows(values_only=True):
                    row_text = ' '.join([str(c) for c in row if c is not None])
                    if row_text.strip():
                        texts.append(row_text)
            result['text'] = '\n'.join(texts)
            wb.close()
        except Exception:
            pass
    
    return result
