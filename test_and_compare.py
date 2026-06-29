#!/usr/bin/env python3
"""测试：处理源文件 → 生成Excel → 对比客户手工参考"""
import sys, os, json
from pathlib import Path
from datetime import datetime
from typing import Dict, Any, List

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from app.pipeline.file_handler import process_file
from app.pipeline.field_extractor import extract_fields, get_field_definitions
from app.pipeline.report_generator import _collect_summary_fields, SUMMARY_HEADERS
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


# ========== 1. 处理所有源文件 ==========

def process_all_sources():
    """处理所有能处理的源文件"""
    results = {}

    # 企业征信PDF
    for pdf_name, rtype in [
        ('test_enterprise2.pdf', 'corporate'),
        ('test_personal2.pdf', 'personal'),
    ]:
        fp = Path(pdf_name)
        if fp.exists():
            print(f'处理: {pdf_name} (type={rtype})')
            file_result = process_file(str(fp))
            fields = extract_fields(rtype, raw_text=file_result['text'])
            results[rtype] = {
                'file': pdf_name,
                'text_len': len(file_result['text']),
                'fields': fields,
                'raw_text': file_result['text'],
            }
            # 打印关键字段
            fd = get_field_definitions(rtype)
            hits = sum(1 for f in fd if fields.get(f['key'], {}).get('value', ''))
            print(f'  → {hits}/{len(fd)} 字段命中')
            for k in ['company_name', 'establish_info', 'legal_person',
                       'total_balance', 'name', 'total_debt']:
                if k in fields:
                    v = fields[k].get('value', '')
                    if v: print(f'    {k}: {v}')

    # 水母报告图片（OCR待安装）
    img_types = {
        'test_invoice.png': '开票',
        'test_tax.png': '纳税',
        'test_supplier.png': '供应商',
        'test_customer.png': '销售客户',
    }
    for img, label in img_types.items():
        if Path(img).exists():
            print(f'待OCR: {img} ({label}) - 需等PaddlePaddle安装')
            results[label] = {'file': img, 'status': 'pending_ocr'}

    return results


# ========== 2. 按客户格式生成Excel ==========

CLIENT_HEADERS = [
    '公司', '高新/深房', '双签', '地址',
    '成立/诉讼/变更税等级关联风险',
    '产品', '社保人数锐减 场地、设备库存',
    '开票纳税 是否增量',
    '企业征信（日期）（增加授信有额度的）',
    '法人征信（日期）',
    '法人配偶征信（日期）',
    '上下游（国央企上市公司、深圳高新以上记录）',
    '备注',
]


def generate_client_excel(results: Dict, output_path: str):
    """按客户13列格式生成Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = '导出数据'
    ws.sheet_view.showGridLines = False

    # 样式
    header_fill = PatternFill('solid', fgColor='D9EAD3')
    header_font = Font(name='Microsoft YaHei', size=10, bold=True)
    body_font = Font(name='Microsoft YaHei', size=9)
    thin = Side(style='thin', color='cccccc')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical='top')

    # 设置列宽
    widths = [30, 8, 8, 35, 30, 25, 15, 30, 35, 35, 30, 40, 15]
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    # 写表头
    for idx, h in enumerate(CLIENT_HEADERS, 1):
        cell = ws.cell(row=1, column=idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    # 冻结首行
    ws.freeze_panes = 'A2'

    # ===== 从结果中提取数据 =====
    corp = results.get('corporate', {})
    fields_c = corp.get('fields', {})
    raw_c = corp.get('raw_text', '')

    pers = results.get('personal', {})
    fields_p = pers.get('fields', {})
    raw_p = pers.get('raw_text', '')

    def v(key, src='corporate'):
        d = fields_c if src == 'corporate' else fields_p
        return (d.get(key, {}) or {}).get('value', '')

    # ===== 使用框架的 _collect_summary_fields 获取核心数据 =====
    corp_summary = _collect_summary_fields('corporate', fields_c, raw_c)
    pers_summary = _collect_summary_fields('personal', fields_p, raw_p)

    import re
    col_company = corp_summary.get('公司高新/深房', '') or v('company_name') or ''

    # 列2: 高新/深房
    col_hightech = ''
    if '高新' in raw_c:
        col_hightech = '高新'

    # 列3: 双签
    col_double = ''

    # 列4: 地址 - 从原文提取
    col_addr = ''
    import re
    m_addr = re.search(r'登记地址[\\n\s]+([^\\n]+)', raw_c)
    if m_addr:
        col_addr = m_addr.group(1).strip().replace(' ', '')
    if not col_addr:
        m_addr2 = re.search(r'办公[\\n\s]*地址[\\n\s]+([^\\n]+)', raw_c)
        if m_addr2:
            col_addr = m_addr2.group(1).strip().replace(' ', '')

    # 列5: 成立/诉讼/变更税等级关联风险
    col_risk = corp_summary.get('成立/诉讼/变更税等级关联风险', '')
    if not col_risk:
        risk_parts = []
        m_e = re.search(r'成立年份\s*(\d{4})', raw_c)
        if m_e: risk_parts.append(f'{m_e.group(1)}年成立')
        m_l = re.search(r'法定代表人[^\\n]*[\\n\s]+(\S+)', raw_c)
        if m_l: risk_parts.append(f'法人{m_l.group(1)}')
        m_t = re.search(r'纳税信用\s*([A-Z]+级)', raw_c)
        if m_t: risk_parts.append(m_t.group(1))
        col_risk = '\n'.join(risk_parts)

    # 列6: 产品
    col_product = ''

    # 列7: 社保
    col_social = ''

    # 列8: 开票纳税（来自 _collect_summary_fields + OCR待补充）
    col_invoice = corp_summary.get('开票纳税', '')
    if not col_invoice:
        col_invoice = '(开票纳税数据在\n近三年开票信息报表.png\n近三年纳税信息完税表.png\n需OCR识别)'

    # 列9: 企业征信（来自 _collect_summary_fields）
    col_enterprise_credit = corp_summary.get('企业征信', '')
    if not col_enterprise_credit:
        balance = v('total_balance')
        col_enterprise_credit = f'负债{balance}万' if balance else ''

    # 列10: 法人征信（来自 _collect_summary_fields + personal）
    pers_summary = _collect_summary_fields('personal', fields_p, raw_p)
    col_personal_credit = corp_summary.get('法人征信', '') or pers_summary.get('法人征信', '')
    if not col_personal_credit:
        name_p = v('name', 'personal')
        if name_p:
            col_personal_credit = f'{name_p}的个人征信数据(待详细提取)'

    # 列11: 法人配偶
    col_spouse = ''

    # 列12: 上下游
    col_chain = ''

    # 列13: 备注
    col_note = ''

    # 写数据行
    row_data = [
        col_company,
        col_hightech,
        col_double,
        col_addr,
        col_risk,
        col_product,
        col_social,
        col_invoice,
        col_enterprise_credit,
        col_personal_credit,
        col_spouse,
        col_chain.strip(),
        col_note,
    ]

    for idx, val in enumerate(row_data, 1):
        cell = ws.cell(row=2, column=idx, value=val or '')
        cell.font = body_font
        cell.alignment = wrap
        cell.border = border

    wb.save(output_path)
    print(f'\n✅ 已生成: {output_path}')
    return output_path


# ========== 3. 对比客户参考 ==========

def compare_with_reference(our_file: str, ref_file: str):
    """对比生成结果和客户手工参考"""
    print(f'\n{"="*60}')
    print(f'📊 对比分析: 生成文件 vs 客户手工参考')
    print(f'{"="*60}')

    our = load_workbook(our_file)
    ref = load_workbook(ref_file)

    our_ws = our.active
    ref_ws = ref.active

    # 对比每一列
    for col in range(1, 14):
        our_val = str(our_ws.cell(row=2, column=col).value or '')
        ref_val = str(ref_ws.cell(row=2, column=col).value or '')
        header = CLIENT_HEADERS[col - 1]

        our_clean = our_val.strip()
        ref_clean = ref_val.strip()

        # 简化对比：检查关键内容是否匹配
        match = '✅' if (our_clean and ref_clean and
                        (our_clean[:20] in ref_clean or ref_clean[:20] in our_clean)) else '⚠️'

        print(f'\n{match} 列{col}: {header}')
        print(f'  生成: {our_clean[:120]}')
        print(f'  客户: {ref_clean[:120]}')

        if not our_clean:
            print(f'  → 待补充')
        elif match == '⚠️':
            print(f'  → 内容不完全匹配，需人工核对')


# ========== 主流程 ==========

if __name__ == '__main__':
    print('=' * 60)
    print('征信报告OCR工具 - 端到端测试 & 对比')
    print('=' * 60)
    print()

    # 1. 处理源文件
    results = process_all_sources()

    # 2. 生成Excel
    output_path = f'output/对比测试_导出_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    generate_client_excel(results, output_path)

    # 3. 对比客户参考
    ref_path = '/Users/apple/Library/Containers/com.tencent.xinWeChat/Data/Documents/xwechat_files/wxid_2xm1gvusd9pn12_86bb/msg/file/2026-06/新建文件夹 (3)/客户资料表格.xlsx'
    if Path(ref_path).exists():
        compare_with_reference(output_path, ref_path)
    else:
        print(f'\n❌ 客户参考文件不存在: {ref_path}')

    print(f'\n{"="*60}')
    print('完成！')
